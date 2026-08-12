#!/usr/bin/env python3
"""Build benchmark tracking workbooks from versioned JSONL history records."""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HISTORY_SCHEMA_VERSION = "memory-ab-history-v1"

HDR_FILL = PatternFill("solid", fgColor="152238")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
COND_FILL = PatternFill("solid", fgColor="E8EDF3")
COND_FONT = Font(bold=True, color="1F3A5F")
METRIC_FILL = PatternFill("solid", fgColor="F5F7FB")
SECTION_FILL = PatternFill("solid", fgColor="CDD9E5")
SECTION_FONT = Font(bold=True, color="152238", size=11)
PCT_FMT = "0.00%"
NUM_FMT = "0.00"
INT_FMT = "#,##0"
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

CONDITION_ROWS = [
    ("Run ID", "run_id"),
    ("Date", "date"),
    ("Backend", "backend"),
    ("Git hash", "git_hash"),
    ("--- Embedding ---", None),
    ("Embedding provider", "embedding_provider"),
    ("Embedding model", "embedding_model"),
    ("Dimensions", "dimensions"),
    ("--- Retrieval ---", None),
    ("Search mode", "search_mode"),
    ("Embedding weight", "embedding_weight"),
    ("BM25 weight", "bm25_weight"),
    ("Store backend", "store_backend"),
    ("Candidate-K (dense/bm25 pool)", "candidate_k"),
    ("Rerank enabled", "rerank_enabled"),
    ("Rerank model", "rerank_model"),
    ("Rerank input-K (post-fuse)", "rerank_input_k"),
    ("Top-K (final return)", "top_k"),
    ("QA top-K", "qa_top_k"),
    ("Context token budget", "context_token_budget"),
    ("--- Models ---", None),
    ("Answer model", "answer_model"),
    ("Judge model", "judge_model"),
    ("Answer prompt version", "answer_prompt_version"),
    ("Memory format", "memory_format"),
    ("--- Notes ---", None),
    ("Notes", "notes"),
    ("--- Memory A/B ---", None),
    ("History schema version", "schema_version"),
    ("Pair ID", "pair_id"),
    ("Dataset", "dataset"),
    ("Split", "split"),
    ("Memory mode", "memory_mode"),
    ("Phase", "phase"),
    ("Source hash", "source_hash"),
    ("Code hash", "code_hash"),
    ("Configuration hash", "configuration_hash"),
    ("Preflight hash", "preflight_hash"),
    ("Policy hash", "policy_hash"),
    ("Promotion status", "promotion_status"),
    ("Promotion reasons", "promotion_reasons"),
    ("Artifact path", "artifact_path"),
]

PERSONALMEM_METRICS = [
    ("QA Accuracy (4-option)", PCT_FMT, ("qa", "overall", "accuracy")),
    (
        "Track Full Preference Evolution",
        PCT_FMT,
        ("qa", "by_question_type", "track_full_preference_evolution", "accuracy"),
    ),
    (
        "Suggest New Ideas",
        PCT_FMT,
        ("qa", "by_question_type", "suggest_new_ideas", "accuracy"),
    ),
    (
        "Recall User Shared Facts",
        PCT_FMT,
        ("qa", "by_question_type", "recall_user_shared_facts", "accuracy"),
    ),
    (
        "Generalizing To New Scenarios",
        PCT_FMT,
        ("qa", "by_question_type", "generalizing_to_new_scenarios", "accuracy"),
    ),
    (
        "Provide Preference Aligned Recommendations",
        PCT_FMT,
        (
            "qa",
            "by_question_type",
            "provide_preference_aligned_recommendations",
            "accuracy",
        ),
    ),
    (
        "Recalling The Reasons Behind Previous Updates",
        PCT_FMT,
        (
            "qa",
            "by_question_type",
            "recalling_the_reasons_behind_previous_updates",
            "accuracy",
        ),
    ),
    (
        "Recalling Facts Mentioned By The User",
        PCT_FMT,
        (
            "qa",
            "by_question_type",
            "recalling_facts_mentioned_by_the_user",
            "accuracy",
        ),
    ),
    ("Avg context tokens", INT_FMT, ("cost", "avg_context_tokens")),
    ("Query count", INT_FMT, ("qa", "overall", "count")),
]

LONGMEMEVAL_METRICS = [
    ("QA Accuracy", PCT_FMT, ("qa", "overall", "accuracy")),
    (
        "Single-session (user)",
        PCT_FMT,
        ("qa", "by_type", "single-session-user", "accuracy"),
    ),
    (
        "Single-session (assistant)",
        PCT_FMT,
        ("qa", "by_type", "single-session-assistant", "accuracy"),
    ),
    (
        "Single-session (preference)",
        PCT_FMT,
        ("qa", "by_type", "single-session-preference", "accuracy"),
    ),
    (
        "Knowledge update",
        PCT_FMT,
        ("qa", "by_type", "knowledge-update", "accuracy"),
    ),
    (
        "Temporal reasoning",
        PCT_FMT,
        ("qa", "by_type", "temporal-reasoning", "accuracy"),
    ),
    ("Multi-session", PCT_FMT, ("qa", "by_type", "multi-session", "accuracy")),
    (
        "Retrieval session R@10",
        PCT_FMT,
        ("retrieval", "session", "overall", "recall_at_10"),
    ),
    (
        "Retrieval turn R@10",
        PCT_FMT,
        ("retrieval", "turn", "overall", "recall_at_10"),
    ),
    (
        "Retrieval turn MRR",
        NUM_FMT,
        ("retrieval", "turn", "overall", "mrr"),
    ),
    ("Avg total tokens", INT_FMT, ("cost", "avg_total_tokens")),
    ("Question count", INT_FMT, ("qa", "overall", "count")),
]

LOCOMO_METRICS = [
    ("LLM Judge Score", PCT_FMT, ("qa", "overall", "llm_score")),
    ("Multi hop (cat1)", PCT_FMT, ("qa", "by_category", "1", "llm_score")),
    (
        "Temporal reasoning (cat2)",
        PCT_FMT,
        ("qa", "by_category", "2", "llm_score"),
    ),
    ("Open domain (cat3)", PCT_FMT, ("qa", "by_category", "3", "llm_score")),
    ("Single hop (cat4)", PCT_FMT, ("qa", "by_category", "4", "llm_score")),
    (
        "Evidence Hit@K",
        PCT_FMT,
        ("retrieval", "overall", "evidence_hit_at_k"),
    ),
    ("Evidence MRR", NUM_FMT, ("retrieval", "overall", "evidence_mrr")),
    ("Avg total tokens", INT_FMT, ("qa", "overall", "avg_total_tokens")),
    ("Question count (excl. cat5)", INT_FMT, ("qa", "overall", "count")),
]

DATASETS = [
    ("personalmem", "PersonaMem", PERSONALMEM_METRICS),
    ("longmemeval", "LongMemEval", LONGMEMEVAL_METRICS),
    ("locomo", "LoCoMo", LOCOMO_METRICS),
]

CONFIGURATION_ALIASES = {
    "embedding_provider": ("embedding_provider", "embedding"),
    "dimensions": ("dimensions", "embedding_dimensions"),
    "top_k": ("top_k", "retrieval_top_k"),
    "answer_model": ("answer_model", "answerer_model", "chat_model"),
}


def load_records(path: Path) -> list[dict[str, Any]]:
    """Load one versioned JSON object per non-empty line from *path*."""
    path = Path(path)
    if not path.is_file():
        return []
    records: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as stream:
        for line_number, line in enumerate(stream, start=1):
            if not line.strip():
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError as error:
                raise ValueError(f"invalid JSONL at {path}:{line_number}") from error
            if not isinstance(record, dict):
                raise ValueError(
                    f"history record at {path}:{line_number} must be an object"
                )
            if record.get("schema_version") != HISTORY_SCHEMA_VERSION:
                raise ValueError(
                    f"history record at {path}:{line_number} must use "
                    f"{HISTORY_SCHEMA_VERSION}"
                )
            records.append(record)
    return records


def append_full_pair(
    record_path: Path, pair_records: Sequence[Mapping[str, Any]]
) -> None:
    """Append one validated completed full raw/extracted pair to JSONL history."""
    records = [dict(record) for record in pair_records]
    if not _is_completed_full_pair(records):
        raise ValueError("append requires a completed full raw/extracted pair")
    serialized = "".join(
        json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        + "\n"
        for record in records
    )
    path = Path(record_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as stream:
        stream.write(serialized)


def _is_completed_full_pair(records: list[dict[str, Any]]) -> bool:
    if len(records) != 2:
        return False
    by_mode = {record.get("memory_mode"): record for record in records}
    if set(by_mode) != {"raw", "extracted"}:
        return False
    raw = by_mode["raw"]
    extracted = by_mode["extracted"]
    required = (
        "pair_id",
        "run_id",
        "dataset",
        "split",
        "source_hash",
        "code_hash",
        "configuration_hash",
        "artifact_path",
    )
    if any(
        record.get("schema_version") != HISTORY_SCHEMA_VERSION
        or record.get("phase") != "full"
        or ("complete" in record and record.get("complete") is not True)
        or any(not record.get(key) for key in required)
        or not isinstance(record.get("configuration"), dict)
        or not isinstance(record.get("metrics"), dict)
        or not isinstance(record.get("promotion_reasons"), list)
        for record in records
    ):
        return False
    governance_mode = raw.get("governance_mode", "strict")
    extracted_governance_mode = extracted.get("governance_mode", "strict")
    if governance_mode not in {"normal", "strict"} or extracted_governance_mode != governance_mode:
        return False
    if governance_mode == "strict" and any(
        not record.get(key) for record in records for key in ("preflight_hash", "policy_hash")
    ):
        return False
    if governance_mode == "normal" and any(
        record.get(key) for record in records for key in ("preflight_hash", "policy_hash")
    ):
        return False
    shared = (
        "pair_id",
        "dataset",
        "split",
        "source_hash",
        "code_hash",
        "configuration_hash",
        "preflight_hash",
        "policy_hash",
    )
    if any(raw[key] != extracted[key] for key in shared):
        return False
    if governance_mode == "normal":
        return (
            raw.get("promotion_status") == "not_evaluated"
            and extracted.get("promotion_status") == "not_evaluated"
            and not raw["promotion_reasons"]
            and not extracted["promotion_reasons"]
        )
    if raw.get("promotion_status") != "reference" or raw["promotion_reasons"]:
        return False
    status = extracted.get("promotion_status")
    reasons = extracted["promotion_reasons"]
    if status not in {"passed", "failed"}:
        return False
    if status == "failed" and (
        not reasons
        or any(not isinstance(reason, str) or not reason for reason in reasons)
    ):
        return False
    if status == "passed" and reasons:
        return False
    return True


def build_workbooks(history_root: Path, records_root: Path | None = None) -> None:
    """Generate the combined and per-dataset workbooks from JSONL records."""
    history_root = Path(history_root)
    source_root = Path(records_root or history_root / "records")
    history_root.mkdir(parents=True, exist_ok=True)
    loaded = {
        stem: load_records(source_root / f"{stem}.jsonl")
        for stem, _, _ in DATASETS
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    for stem, sheet_name, metrics in DATASETS:
        _build_sheet(workbook, sheet_name, loaded[stem], metrics)
    _save_atomic(workbook, history_root / "benchmarks.xlsx")

    for stem, sheet_name, metrics in DATASETS:
        workbook = Workbook()
        workbook.remove(workbook.active)
        _build_sheet(workbook, sheet_name, loaded[stem], metrics)
        _save_atomic(workbook, history_root / f"{stem}_tracking.xlsx")


def _save_atomic(workbook: Workbook, destination: Path) -> None:
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    try:
        workbook.save(temporary)
    finally:
        workbook.close()
    os.replace(temporary, destination)


def _build_sheet(
    workbook: Workbook,
    dataset_name: str,
    records: list[dict[str, Any]],
    metrics_spec: list[tuple[str, str, tuple[str, ...]]],
    n_future_cols: int = 5,
):
    worksheet = workbook.create_sheet(title=dataset_name)
    n_existing = len(records)
    n_columns = 1 + n_existing + n_future_cols

    worksheet.cell(1, 1, value=f"{dataset_name} — benchmark tracking")
    worksheet.cell(1, 1).font = Font(bold=True, size=14, color="152238")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_columns)
    worksheet.row_dimensions[1].height = 22

    worksheet.cell(3, 1, value="Metric / Condition")
    for index, record in enumerate(records, start=2):
        worksheet.cell(3, index, value=record["run_id"])
    for offset in range(n_future_cols):
        worksheet.cell(3, 2 + n_existing + offset, value=f"(future run {offset + 1})")
    _style_header_row(worksheet, 3, n_columns)
    worksheet.row_dimensions[3].height = 30

    row = _write_condition_block(worksheet, 4, records, n_columns)
    row = _write_section(worksheet, row, "=== Metrics ===", n_columns)
    row = _write_metric_block(worksheet, row, metrics_spec, records)
    _add_blank_run_columns(worksheet, 2 + n_existing, n_columns, row)

    worksheet.column_dimensions["A"].width = 34
    for column in range(2, n_columns + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 16
    worksheet.freeze_panes = "B4"
    return worksheet


def _style_header_row(worksheet, row: int, n_columns: int) -> None:
    for column in range(1, n_columns + 1):
        cell = worksheet.cell(row, column)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER


def _write_section(worksheet, row: int, text: str, n_columns: int) -> int:
    worksheet.cell(row, 1, value=text)
    for column in range(1, n_columns + 1):
        cell = worksheet.cell(row, column)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    worksheet.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=n_columns
    )
    return row + 1


def _write_condition_block(
    worksheet,
    start_row: int,
    records: list[dict[str, Any]],
    n_columns: int,
) -> int:
    row = start_row
    for label, key in CONDITION_ROWS:
        if key is None:
            row = _write_section(worksheet, row, label, n_columns)
            continue
        cell = worksheet.cell(row, 1, value=label)
        cell.fill = COND_FILL
        cell.font = COND_FONT
        cell.alignment = LEFT
        cell.border = BORDER
        for index, record in enumerate(records, start=2):
            value = _condition_value(record, key)
            record_cell = worksheet.cell(row, index, value=value)
            record_cell.alignment = LEFT
            record_cell.border = BORDER
        row += 1
    return row


def _condition_value(record: dict[str, Any], key: str) -> Any:
    if key in record:
        value = record[key]
    else:
        configuration = record.get("configuration")
        value = ""
        if isinstance(configuration, dict):
            aliases = CONFIGURATION_ALIASES.get(key, (key,))
            for alias in aliases:
                if alias in configuration:
                    value = configuration[alias]
                    break
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_metric_block(
    worksheet,
    start_row: int,
    metrics_spec: list[tuple[str, str, tuple[str, ...]]],
    records: list[dict[str, Any]],
) -> int:
    row = start_row
    for label, number_format, path in metrics_spec:
        cell = worksheet.cell(row, 1, value=label)
        cell.fill = METRIC_FILL
        cell.font = Font(bold=True)
        cell.alignment = LEFT
        cell.border = BORDER
        for index, record in enumerate(records, start=2):
            value = _metric_value(record, label, path)
            metric_cell = worksheet.cell(row, index)
            if value is not None and value != "":
                metric_cell.value = value
                metric_cell.number_format = number_format
            metric_cell.alignment = RIGHT
            metric_cell.border = BORDER
        row += 1
    return row


def _metric_value(
    record: dict[str, Any], label: str, path: tuple[str, ...]
) -> Any:
    metrics = record.get("metrics")
    if not isinstance(metrics, dict):
        return None
    if label in metrics:
        return metrics[label]
    value: Any = metrics
    for segment in path:
        if not isinstance(value, dict) or segment not in value:
            return None
        value = value[segment]
    return value


def _add_blank_run_columns(
    worksheet, start_column: int, end_column: int, n_rows: int
) -> None:
    for column in range(start_column, end_column + 1):
        for row in range(1, n_rows + 1):
            cell = worksheet.cell(row, column)
            if not cell.border:
                cell.border = BORDER


def main() -> None:
    history_root = Path(__file__).resolve().parent
    build_workbooks(history_root)
    for filename in (
        "benchmarks.xlsx",
        "personalmem_tracking.xlsx",
        "longmemeval_tracking.xlsx",
        "locomo_tracking.xlsx",
    ):
        print(f"wrote {history_root / filename}")


if __name__ == "__main__":
    main()
