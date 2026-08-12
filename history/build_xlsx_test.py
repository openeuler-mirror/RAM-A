from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

import history.build_xlsx as build_xlsx
from history.build_xlsx import append_full_pair, build_workbooks, load_records


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RECORDS_ROOT = PROJECT_ROOT / "history" / "records"
EVALUATION_REQUIREMENTS = PROJECT_ROOT / "evaluation" / "requirements.txt"

EXPECTED_RUN_IDS = {
    "personalmem": [
        "v1 baseline\n(memory-euler)\n2026-06-10",
        "v2 bm25+sqlite\n2026-07-03",
        "v2' bm25+sqlite\ntop_k=10\n2026-07-08",
        "v3 +rerank\ntop_k=20\n2026-07-08",
    ],
    "longmemeval": [
        "v1 baseline\n(memory-euler)\n2026-06-03",
        "v2 bm25+sqlite\n2026-07-03",
        "v3 +rerank\ntop_k=10\n2026-07-09",
    ],
    "locomo": [
        "v1 baseline\n(memory-euler)\n2026-06-10",
        "v2 bm25+sqlite\n2026-07-07",
        "v3 +rerank\ncohere-v3.5\n2026-07-08",
        "v4 A/B fresh raw\n+rerank\n2026-07-16",
        "v5 A/B extracted\n+rerank\n2026-07-16",
        "locomo-raw-59ba28e03223e14e",
        "locomo-extracted-0756fa66cf3bb0a6",
    ],
}

EXISTING_CONDITION_LABELS = [
    "Run ID",
    "Date",
    "Backend",
    "Git hash",
    "--- Embedding ---",
    "Embedding provider",
    "Embedding model",
    "Dimensions",
    "--- Retrieval ---",
    "Search mode",
    "Embedding weight",
    "BM25 weight",
    "Store backend",
    "Candidate-K (dense/bm25 pool)",
    "Rerank enabled",
    "Rerank model",
    "Rerank input-K (post-fuse)",
    "Top-K (final return)",
    "QA top-K",
    "Context token budget",
    "--- Models ---",
    "Answer model",
    "Judge model",
    "Answer prompt version",
    "Memory format",
    "--- Notes ---",
    "Notes",
]


def _history_record(
    memory_mode: str,
    *,
    promotion_status: str,
    promotion_reasons: list[str] | None = None,
) -> dict[str, object]:
    return {
        "schema_version": "memory-ab-history-v1",
        "pair_id": "pair-1",
        "run_id": f"run-{memory_mode}",
        "dataset": "personalmem",
        "split": "32k",
        "memory_mode": memory_mode,
        "phase": "full",
        "source_hash": "source-sha",
        "code_hash": "code-sha",
        "configuration_hash": "config-sha",
        "preflight_hash": "preflight-sha",
        "policy_hash": "policy-sha",
        "configuration": {"top_k": 10, "answer_model": "answer-model"},
        "metrics": {"qa": {"overall": {"accuracy": 0.75, "count": 589}}},
        "promotion_status": promotion_status,
        "promotion_reasons": promotion_reasons or [],
        "artifact_path": f"/artifacts/{memory_mode}",
    }


def _row_for_label(worksheet, label: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == label:
            return row
    raise AssertionError(f"missing row label: {label}")


def test_migrated_jsonl_preserves_every_existing_run_in_order() -> None:
    records_by_dataset = {
        name: load_records(RECORDS_ROOT / f"{name}.jsonl")
        for name in EXPECTED_RUN_IDS
    }

    assert {
        name: [record["run_id"] for record in records]
        for name, records in records_by_dataset.items()
    } == EXPECTED_RUN_IDS
    assert all(
        record["schema_version"] == "memory-ab-history-v1"
        for records in records_by_dataset.values()
        for record in records
    )
    assert records_by_dataset["personalmem"][0]["pair_id"] == ""
    assert records_by_dataset["longmemeval"][0]["promotion_status"] == ""
    assert records_by_dataset["locomo"][3]["metrics"]["Avg total tokens"] == 1760.18
    assert records_by_dataset["locomo"][4]["metrics"]["LLM Judge Score"] == 0.5682


def test_evaluation_requirements_declare_openpyxl_for_history_workbook_generation() -> None:
    declared_requirements = [
        line.strip().lower()
        for line in EVALUATION_REQUIREMENTS.read_text().splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    ]

    assert any(requirement.startswith("openpyxl>=") for requirement in declared_requirements)


def test_existing_history_round_trips_to_expected_columns(tmp_path: Path) -> None:
    build_workbooks(tmp_path, records_root=RECORDS_ROOT)

    workbook = load_workbook(tmp_path / "benchmarks.xlsx", data_only=True)
    assert workbook.sheetnames == ["PersonaMem", "LongMemEval", "LoCoMo"]
    assert workbook["LoCoMo"]["E3"].value == (
        "v4 A/B fresh raw\n+rerank\n2026-07-16"
    )
    assert workbook["LoCoMo"]["F3"].value == (
        "v5 A/B extracted\n+rerank\n2026-07-16"
    )
    assert [
        workbook["PersonaMem"].cell(3, column).value for column in range(2, 6)
    ] == EXPECTED_RUN_IDS["personalmem"]
    assert [
        workbook["LongMemEval"].cell(3, column).value for column in range(2, 5)
    ] == EXPECTED_RUN_IDS["longmemeval"]


def test_build_workbooks_allows_missing_dataset_history_files(tmp_path: Path) -> None:
    records_root = tmp_path / "records"
    records_root.mkdir()
    source = RECORDS_ROOT / "locomo.jsonl"
    (records_root / source.name).write_bytes(source.read_bytes())

    build_workbooks(tmp_path / "workbooks", records_root=records_root)

    assert (tmp_path / "workbooks" / "benchmarks.xlsx").is_file()


def test_task6_nested_records_populate_ab_conditions_and_locomo_metrics(
    tmp_path: Path,
) -> None:
    records_root = tmp_path / "records"
    records_root.mkdir()
    for source in RECORDS_ROOT.glob("*.jsonl"):
        (records_root / source.name).write_bytes(source.read_bytes())
    pair = [
        _history_record("raw", promotion_status="reference"),
        _history_record("extracted", promotion_status="passed"),
    ]
    for index, record in enumerate(pair):
        record["dataset"] = "locomo"
        record["configuration"] = {
            "chat_model": "answer-model",
            "top_k": 30,
        }
        record["metrics"] = {
            "qa": {
                "overall": {
                    "llm_score": 0.5 + index / 10,
                    "count": 1540,
                    "avg_total_tokens": 3000 + index,
                },
                "by_category": {},
            },
            "retrieval": {
                "overall": {
                    "evidence_hit_at_k": 0.4 + index / 10,
                    "evidence_mrr": 0.2 + index / 10,
                }
            },
        }
    append_full_pair(records_root / "locomo.jsonl", pair)

    build_workbooks(tmp_path / "workbooks", records_root=records_root)

    worksheet = load_workbook(tmp_path / "workbooks" / "benchmarks.xlsx")["LoCoMo"]
    extracted_column = next(
        column
        for column in range(2, worksheet.max_column + 1)
        if worksheet.cell(3, column).value == "run-extracted"
    )
    expected = {
        "Pair ID": "pair-1",
        "Memory mode": "extracted",
        "Promotion status": "passed",
        "Answer model": "answer-model",
        "LLM Judge Score": 0.6,
        "Evidence Hit@K": 0.5,
        "Avg total tokens": 3001,
    }
    assert {
        label: worksheet.cell(_row_for_label(worksheet, label), extracted_column).value
        for label in expected
    } == expected


@pytest.mark.parametrize(
    ("sheet_name", "existing_count", "metric_label", "number_format"),
    [
        ("PersonaMem", 4, "QA Accuracy (4-option)", "0.00%"),
        ("LongMemEval", 3, "Retrieval turn MRR", "0.00"),
        ("LoCoMo", 7, "Question count (excl. cat5)", "#,##0"),
    ],
)
def test_existing_workbook_layout_styles_and_future_columns_are_preserved(
    tmp_path: Path,
    sheet_name: str,
    existing_count: int,
    metric_label: str,
    number_format: str,
) -> None:
    build_workbooks(tmp_path, records_root=RECORDS_ROOT)
    worksheet = load_workbook(tmp_path / "benchmarks.xlsx")[sheet_name]

    labels = [worksheet.cell(row, 1).value for row in range(4, worksheet.max_row + 1)]
    assert labels[: len(EXISTING_CONDITION_LABELS)] == EXISTING_CONDITION_LABELS
    assert "--- Memory A/B ---" in labels
    assert "Pair ID" in labels
    assert "Memory mode" in labels
    assert "Promotion status" in labels
    assert worksheet.freeze_panes == "B4"
    assert worksheet.column_dimensions["A"].width == 34
    assert all(
        worksheet.column_dimensions[letter].width == 16
        for letter in "BCDEFGHIJKLMNOP"[: existing_count + 5]
    )
    assert [
        worksheet.cell(3, column).value
        for column in range(2 + existing_count, 2 + existing_count + 5)
    ] == [f"(future run {number})" for number in range(1, 6)]
    metric_row = _row_for_label(worksheet, metric_label)
    assert worksheet.cell(metric_row, 2).number_format == number_format
    assert worksheet["A1"].font.bold is True
    assert worksheet["A3"].fill.fgColor.rgb == "00152238"


def test_builds_atomic_combined_and_per_dataset_workbooks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    real_replace = os.replace
    replacements: list[tuple[Path, Path]] = []

    def recording_replace(source, destination) -> None:
        source_path = Path(source)
        destination_path = Path(destination)
        assert source_path == destination_path.with_suffix(
            destination_path.suffix + ".tmp"
        )
        assert source_path.exists()
        replacements.append((source_path, destination_path))
        real_replace(source_path, destination_path)

    monkeypatch.setattr(build_xlsx.os, "replace", recording_replace)

    build_workbooks(tmp_path, records_root=RECORDS_ROOT)

    assert {destination.name for _, destination in replacements} == {
        "benchmarks.xlsx",
        "personalmem_tracking.xlsx",
        "longmemeval_tracking.xlsx",
        "locomo_tracking.xlsx",
    }
    assert not list(tmp_path.glob("*.tmp"))


def test_main_reports_the_actual_personalmem_workbook_name(
    monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    monkeypatch.setattr(build_xlsx, "build_workbooks", lambda history_root: None)

    build_xlsx.main()

    output = capsys.readouterr().out
    assert "personalmem_tracking.xlsx" in output
    assert "personamem_tracking.xlsx" not in output


@pytest.mark.parametrize(
    "pair_records",
    [
        [{"phase": "pilot", "memory_mode": "raw"}],
        [
            {"phase": "full", "memory_mode": "raw"},
            {"phase": "full", "memory_mode": "raw"},
        ],
        [
            {"phase": "full", "memory_mode": "raw", "complete": False},
            {"phase": "full", "memory_mode": "extracted", "complete": False},
        ],
    ],
)
def test_append_rejects_incomplete_or_non_full_pair_without_writing(
    tmp_path: Path, pair_records: list[dict[str, object]]
) -> None:
    record_path = tmp_path / "p.jsonl"

    with pytest.raises(ValueError, match="completed full raw/extracted pair"):
        append_full_pair(record_path, pair_records)

    assert not record_path.exists()


def test_append_records_completed_failed_promotion_pair_without_baselining_it(
    tmp_path: Path,
) -> None:
    record_path = tmp_path / "p.jsonl"
    pair = [
        _history_record("raw", promotion_status="reference"),
        _history_record(
            "extracted",
            promotion_status="failed",
            promotion_reasons=["historical_primary"],
        ),
    ]

    append_full_pair(record_path, pair)

    stored = load_records(record_path)
    assert stored == pair
    assert stored[1]["promotion_status"] == "failed"
    assert stored[1]["promotion_status"] != "passed"


def test_load_records_rejects_unversioned_jsonl(tmp_path: Path) -> None:
    path = tmp_path / "records.jsonl"
    path.write_text(json.dumps({"run_id": "legacy"}) + "\n", encoding="utf-8")

    with pytest.raises(ValueError, match="memory-ab-history-v1"):
        load_records(path)


def test_gitignore_keeps_sources_trackable_and_generated_files_ignored() -> None:
    rules = (PROJECT_ROOT / ".gitignore").read_text(encoding="utf-8").splitlines()

    assert "/history/" not in rules
    assert "/history/*.xlsx" in rules
    assert "/history/outputs/" in rules
    assert "/history/*.xlsx.tmp" in rules
    assert "/history/.~lock.*#" in rules
